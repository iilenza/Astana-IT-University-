import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

# ------------------- Пути -------------------
data_path = r"C:\Users\а\Downloads\DV_ASS2\datasets"
charts_path = r"C:\Users\а\Downloads\DV_ASS2\charts"
exports_path = r"C:\Users\а\Downloads\DV_ASS2\exports"
os.makedirs(charts_path, exist_ok=True)
os.makedirs(exports_path, exist_ok=True)

# ------------------- Загрузка данных -------------------
students = pd.read_csv(os.path.join(data_path, "students_raw.csv"))
grades = pd.read_csv(os.path.join(data_path, "grades_raw.csv"))
enrollment = pd.read_csv(os.path.join(data_path, "enrollment_raw.csv"))
attendance = pd.read_csv(os.path.join(data_path, "attendance.csv"))

# Очистка email и курсов для join
for df, col in [(students, "email"), (grades, "Email"), (enrollment, "email"), (attendance, "email")]:
    df[col] = df[col].str.lower().str.strip()

enrollment['course_id'] = enrollment['course_id'].astype(str).str.strip()
grades['Course'] = grades['Course'].astype(str).str.strip()
attendance['course'] = attendance['course'].astype(str).str.strip()

# ------------------- JOINs -------------------
# JOIN 1: students + grades
df = students.merge(grades, left_on='email', right_on='Email', how='inner')
# JOIN 2: объединяем с enrollment
df = df.merge(enrollment, on='email', how='left')
# JOIN 3: объединяем с attendance
df = df.merge(attendance, left_on=['email', 'course_id'], right_on=['email', 'course'], how='left')

# Преобразуем числовые поля
df['Total'] = pd.to_numeric(df['Total'], errors='coerce')
df['Midterm'] = pd.to_numeric(df['Midterm'], errors='coerce')
df['attendance'] = pd.to_numeric(df['attendance'], errors='coerce')

print("Размер объединённого датасета:", df.shape)

# ------------------- Агрегированные данные для Excel -------------------
agg_stepen = df['stepen'].value_counts().reset_index()
agg_stepen.columns = ['Stepen', 'Count']

agg_avg_total_course = df.groupby('Course')['Total'].mean().sort_values(ascending=False).head(10).reset_index()

agg_funding = df.groupby('funding')['email'].nunique().sort_values(ascending=False).reset_index()
agg_funding.columns = ['Funding', 'Count']

agg_avg_midterm = df.groupby('Course')['Midterm'].mean().sort_values(ascending=False).head(5).reset_index()

agg_avg_attendance = df.groupby('Course')['attendance'].mean().sort_values(ascending=False).head(5).reset_index()

# ------------------- Excel экспорт с цветной градацией -------------------
def export_to_excel(dataframes_dict, filename):
    filepath = os.path.join(exports_path, filename)
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for sheet_name, df_export in dataframes_dict.items():
            df_export.to_excel(writer, sheet_name=sheet_name, index=False)
    wb = load_workbook(filepath)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = "B2"
        for col in range(2, ws.max_column + 1):
            col_letter = ws.cell(row=1, column=col).column_letter
            rule = ColorScaleRule(
                start_type="min", start_color="FFAA0000",
                mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                end_type="max", end_color="FF00AA00"
            )
            ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)
        ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)
    print(f"Создан файл {filename}, {len(dataframes_dict)} лист(ов), всего строк: {sum(len(df) for df in dataframes_dict.values())}")

export_to_excel({
    "By_Stepen": agg_stepen,
    "Avg_Total_Course": agg_avg_total_course,
    "By_Funding": agg_funding,
    "Avg_Midterm": agg_avg_midterm,
    "Avg_Attendance": agg_avg_attendance
}, "students_report_aggregated.xlsx")

# ------------------- 6 графиков Matplotlib -------------------

# 1. Pie chart: распределение студентов по степеням
plt.figure(figsize=(6,6))
colors = plt.cm.Pastel1(range(len(agg_stepen)))
plt.pie(agg_stepen['Count'], labels=agg_stepen['Stepen'], autopct='%1.1f%%', colors=colors, startangle=90, pctdistance=0.8)
plt.title("Распределение студентов по степеням")
plt.axis('equal')
plt.tight_layout()
plt.savefig(os.path.join(charts_path, 'pie_stepen.png'))
plt.close()

# 2. Bar chart: средний Total по ТОП-10 курсам
plt.figure(figsize=(10,6))
plt.bar(agg_avg_total_course['Course'], agg_avg_total_course['Total'], color='skyblue')
plt.title("Средний итоговый балл по ТОП-10 курсам")
plt.ylabel("Средний балл")
plt.xticks(rotation=30, ha='right')
plt.tight_layout()
plt.savefig(os.path.join(charts_path, 'bar_avg_total_by_course.png'))
plt.close()

# 3. Horizontal bar: количество студентов по источнику финансирования
plt.figure(figsize=(8,6))
plt.barh(agg_funding['Funding'], agg_funding['Count'], color='lightgreen')
plt.title("Количество студентов по источнику финансирования")
plt.xlabel("Количество студентов")
plt.tight_layout()
plt.savefig(os.path.join(charts_path, 'hbar_funding.png'))
plt.close()

# 4. Line chart: средний Midterm по ТОП-5 курсам
plt.figure(figsize=(10,6))
plt.plot(agg_avg_midterm['Course'], agg_avg_midterm['Midterm'], marker='o', color='orange')
plt.title("Средний Midterm по ТОП-5 курсам")
plt.ylabel("Средний Midterm")
plt.xticks(rotation=30, ha='right')
plt.tight_layout()
plt.savefig(os.path.join(charts_path, 'line_midterm.png'))
plt.close()

# 5. Histogram: распределение итоговых баллов
plt.figure(figsize=(8,6))
df['Total'].plot.hist(bins=15, edgecolor='black', color='purple')
plt.title("Распределение итоговых баллов")
plt.xlabel("Баллы")
plt.tight_layout()
plt.savefig(os.path.join(charts_path, 'hist_total.png'))
plt.close()

# 6. Scatter plot: посещаемость vs итоговый балл
df_scatter = df[df['attendance'].notnull()]
plt.figure(figsize=(8,6))
plt.scatter(df_scatter['attendance'], df_scatter['Total'], alpha=0.6, color='red')
plt.title("Посещаемость vs Итоговый балл")
plt.xlabel("Посещаемость (%)")
plt.ylabel("Итоговый балл")
plt.tight_layout()
plt.savefig(os.path.join(charts_path, 'scatter_attendance_total.png'))
plt.close()

print("Все 6 графиков Matplotlib созданы в папке charts/")

# ------------------- Plotly интерактивный график (анимация по funding) -------------------
# Сгруппируем данные: средний Total по курсам и источнику финансирования
df_plotly = df.groupby(['Course', 'funding'], as_index=False)['Total'].mean()

# Ограничим до ТОП-10 курсов по среднему Total для наглядности
top_courses = df_plotly.groupby('Course')['Total'].mean().sort_values(ascending=False).head(10).index
df_plotly = df_plotly[df_plotly['Course'].isin(top_courses)]

fig = px.scatter(df_plotly,
                 x="Course",
                 y="Total",
                 size="Total",
                 color="funding",
                 animation_frame="funding",
                 title="Средний итоговый балл по ТОП-10 курсам в зависимости от источника финансирования",
                 range_y=[0, df_plotly['Total'].max() + 10])

fig.update_layout(xaxis_tickangle=-30)
fig.show()

